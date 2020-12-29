from collections import defaultdict, Counter
from openpyxl import load_workbook, worksheet
import datetime 
import pandas
import openpyxl
import collections


# Функция группировки по повторяющимся элементам словаря в словаре
def quantity_in_month(popular):
    for k,v in popular.items():
        popular[k] = Counter(popular[k])
    return popular

# Функция добавления данных в report.xlsx
def data_to_table(element_sort, element_month, j=5, r=12):
    for i in range(1,13):
            sheet.cell(row=r, column=2+i).value = 0
    for i in element_sort: # 7 самых популярных браузеров
        sheet.cell(row=j, column=1).value = i[0]
        for k,v in element_month[i[0]].items():
            sheet.cell(row=j, column=2+k).value = v
            sum = sheet.cell(row=r, column=2+k).value
            sheet.cell(row=r, column=2+k).value = sum + v         
        j +=1


# Чтение данных из фаила logs.xlsx
exel_logs = pandas.read_excel('logs.xlsx', engine='openpyxl', sheet_name='log')
exel_logs_dict = exel_logs.to_dict(orient = 'records')

# Получение списков
browser = defaultdict(int) # Список браузеров по количеству посещений 
browser_month = defaultdict(list) # Список браузеров по количеству посещений по месяцам
item = defaultdict(int) # Список товаров
item_month = defaultdict(list) # Cписок товаров по месяцам
people_items = defaultdict(list) # Список товаров купленных мужчинами и женщинами

for i in exel_logs_dict:
    browser[i['Браузер']] += 1
    browser_month[i['Браузер']].append(i['Дата посещения'].month)
    for k in i['Купленные товары'].split(','):
        item[k] +=1
        item_month[k].append(i['Дата посещения'].month)
    people_items[i['Пол']].extend(i['Купленные товары'].split(','))
    
m_items = Counter(people_items['м']).most_common(len(people_items['м'])) # Мужские товары сгрупированные
w_items = Counter(people_items['ж']).most_common(len(people_items['ж'])) # Женские товары сгрупированные

popular_browser_sort = Counter(browser).most_common(7) # 7 популярных браузеров
popular_item_sort = Counter(item).most_common(7) # 7 популярных товаров

browser_month = quantity_in_month(browser_month) # Определение количества посещений в месяце
item_month = quantity_in_month(item_month) # Определение количества покупок в месяце

# Запись данных в report.xlsx
try:
    wb = load_workbook(filename='report.xlsx')
    sheet = wb['Лист1']
    data_to_table(popular_browser_sort, browser_month, 5, 12) # 7 самых популярных браузеров
    data_to_table(popular_item_sort, item_month, 19, 26) # 7 самых популярных товаров
    sheet.cell(row=31, column=2).value = m_items[0][0] # Самый популярный товар среди мужчин
    sheet.cell(row=32, column=2).value = w_items[0][0] # Самый популярный товар среди женщин
    sheet.cell(row=33, column=2).value = m_items[len(m_items)-1][0] # Самый невостребованный товар среди мужчин
    sheet.cell(row=34, column=2).value = w_items[len(m_items)-1][0] # Самый невостребованный товар среди женщин
    sheet.cell(row=35, column=1).value = 'Самый популярный товар'
    sheet.cell(row=36, column=1).value = 'Самый невостребованный товар'
    sheet.cell(row=35, column=2).value = (Counter(item).most_common(len(item)))[0][0] # Самый популярный товар
    sheet.cell(row=36, column=2).value = (Counter(item).most_common(len(item)))[len(item)-1][0] # Самый невостребованный товар
    wb.save(filename='report.xlsx')
except PermissionError:
    print("Закройте exel фаил report.xlsx")
